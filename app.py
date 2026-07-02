import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from flask import Flask, render_template, request, jsonify, send_file

app = Flask(__name__)

BASE = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE, 'uploads')
OUTPUT_FOLDER = os.path.join(BASE, 'outputs')

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

COLORS = {
    "New Lead / Eligible":    "C6EFCE",
    "Attempted Contact":      "FFEB9C",
    "Screened":               "BDD7EE",
    "Enrolled":               "E2EFDA",
    "Ineligible":             "FFC7CE",
    "New Lead / Ineligible":  "FFD9B3",
}

thin = Side(style="thin", color="DDDDDD")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

SHEET_MAP = {
    "New Lead / Eligible":    "Potential Participants",
    "New Lead / Ineligible":  "Ineligible",
    "Attempted Contact":      "Potential Participants",
    "Screened":               "Potential Participants",
    "Enrolled":               "Scheduled + Active",
    "Ineligible":             "Ineligible",
}

def normalize_phone(phone):
    if pd.isna(phone):
        return ""
    return "".join(filter(str.isdigit, str(phone)))

def get_existing_phones(wb):
    existing = set()
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            for cell in row:
                if cell and isinstance(cell, str) and len(normalize_phone(cell)) >= 10:
                    existing.add(normalize_phone(cell))
    return existing

def find_next_empty_row(ws, start_row=2):
    for row_idx in range(start_row, ws.max_row + 10):
        row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, 6)]
        if all(v is None for v in row_vals):
            return row_idx
    return ws.max_row + 1

def build_row(record):
    name = f"{record.get('First Name', '')} {record.get('Last Name', '')}".strip()
    return [
        None, None,
        str(record.get("Status", "")),
        name,
        str(record.get("Phone Number", "")),
        str(record.get("Email Address", "")),
        "BuildClinical",
        str(record.get("Notes", "")),
        str(record.get("Submission Date", "")),
        str(record.get("Contact Attempts", "")),
        None, None, None, None
    ]

def write_row(ws, row_data, status, row_idx):
    color = COLORS.get(status, "FFFFFF")
    fill = PatternFill("solid", start_color=color, end_color=color)
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = fill
        cell.border = BORDER
        cell.font = Font(name="Calibri", size=9)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row_idx].height = 18

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/run", methods=["POST"])
def run():
    csv_file   = request.files.get("csv")
    excel_file = request.files.get("excel")
    overwrite  = request.form.get("overwrite") == "true"

    if not csv_file or not excel_file:
        return jsonify({"error": "Both files required"}), 400

    csv_path    = os.path.join(UPLOAD_FOLDER, "leads.csv")
    excel_path  = os.path.join(UPLOAD_FOLDER, "tracker.xlsx")
    output_path = os.path.join(OUTPUT_FOLDER, "tracker_updated.xlsx")

    csv_file.save(csv_path)
    excel_file.save(excel_path)

    df = pd.read_csv(csv_path)
    wb = load_workbook(excel_path)
    existing_phones = get_existing_phones(wb)

    added = []
    duplicates = []
    skipped = []

    for _, record in df.iterrows():
        phone_norm = normalize_phone(record.get("Phone Number", ""))
        status = str(record.get("Status", "")).strip()
        name = f"{record.get('First Name', '')} {record.get('Last Name', '')}".strip()

        if phone_norm and phone_norm in existing_phones:
            duplicates.append({"name": name, "phone": str(record.get("Phone Number", ""))})
            continue

        target_sheet = SHEET_MAP.get(status, "Potential Participants")
        if target_sheet not in wb.sheetnames:
            skipped.append({"name": name, "reason": f"Sheet not found for status '{status}'"})
            continue

        ws = wb[target_sheet]
        next_row = find_next_empty_row(ws)
        row_data = build_row(record)
        write_row(ws, row_data, status, next_row)

        if phone_norm:
            existing_phones.add(phone_norm)

        added.append({
            "name": name,
            "sheet": target_sheet,
            "status": status,
            "row": next_row
        })

    # Always save to outputs folder
    wb.save(output_path)

    # If overwrite, also save back to the uploaded tracker path
    if overwrite:
        wb.save(excel_path)

    return jsonify({
        "added": added,
        "duplicates": duplicates,
        "skipped": skipped,
        "total": len(df),
        "overwrite": overwrite
    })

@app.route("/download")
def download():
    path = os.path.join(OUTPUT_FOLDER, "tracker_updated.xlsx")
    return send_file(path, as_attachment=True, download_name="VAREN_Tracker_Updated.xlsx")

if __name__ == "__main__":
    app.run(debug=True, port=5050)
